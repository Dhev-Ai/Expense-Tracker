"""
Enhanced Report View
Advanced analytics with interactive charts, graphs, and export options
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
from datetime import datetime, timedelta

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.styles import COLORS, FONTS, CHART_COLORS
from utils.helpers import format_currency, get_month_name, get_month_short_name
from controllers.expense_controller import ExpenseController

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Try to import matplotlib
try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class ReportView(tk.Frame):
    """Enhanced Report View with interactive charts"""
    
    def __init__(self, parent, user, on_navigate):
        super().__init__(parent, bg=COLORS['bg_secondary'])
        self.parent = parent
        self.user = user
        self.on_navigate = on_navigate
        self.current_period = 'month'  # month, year, custom
        self.current_chart = 'pie'  # pie, bar, line
        
        self.create_widgets()
        self.load_data()
    
    def create_widgets(self):
        """Create report widgets"""
        # Main container with scroll
        canvas = tk.Canvas(self, bg=COLORS['bg_secondary'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg=COLORS['bg_secondary'])
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Bind mouse wheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Content
        content = tk.Frame(self.scrollable_frame, bg=COLORS['bg_secondary'])
        content.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Header with controls
        self.create_header(content)
        
        # Summary cards
        self.summary_frame = tk.Frame(content, bg=COLORS['bg_secondary'])
        self.summary_frame.pack(fill=tk.X, pady=(20, 0))
        
        # Main chart area
        chart_container = tk.Frame(content, bg=COLORS['bg_secondary'])
        chart_container.pack(fill=tk.BOTH, expand=True, pady=(20, 0))
        
        # Left - Main chart
        self.main_chart_frame = tk.Frame(chart_container, bg=COLORS['card_bg'])
        self.main_chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Right - Category breakdown
        self.breakdown_frame = tk.Frame(chart_container, bg=COLORS['card_bg'], width=350)
        self.breakdown_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.breakdown_frame.pack_propagate(False)
        
        # Bottom charts row
        bottom_charts = tk.Frame(content, bg=COLORS['bg_secondary'])
        bottom_charts.pack(fill=tk.X, pady=(20, 0))
        
        # Daily trend chart
        self.daily_chart_frame = tk.Frame(bottom_charts, bg=COLORS['card_bg'])
        self.daily_chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Monthly comparison chart
        self.monthly_chart_frame = tk.Frame(bottom_charts, bg=COLORS['card_bg'])
        self.monthly_chart_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        # Expense table
        self.table_frame = tk.Frame(content, bg=COLORS['card_bg'])
        self.table_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))
    
    def create_header(self, parent):
        """Create header with controls"""
        header = tk.Frame(parent, bg=COLORS['bg_secondary'])
        header.pack(fill=tk.X)
        
        # Title
        left = tk.Frame(header, bg=COLORS['bg_secondary'])
        left.pack(side=tk.LEFT)
        
        tk.Label(
            left,
            text="📊 Analytics & Reports",
            font=FONTS['heading'],
            bg=COLORS['bg_secondary'],
            fg=COLORS['text_primary']
        ).pack(anchor='w')
        
        tk.Label(
            left,
            text="Detailed insights into your spending patterns",
            font=FONTS['body'],
            bg=COLORS['bg_secondary'],
            fg=COLORS['text_secondary']
        ).pack(anchor='w')
        
        # Controls
        right = tk.Frame(header, bg=COLORS['bg_secondary'])
        right.pack(side=tk.RIGHT)
        
        # Period selector
        period_frame = tk.Frame(right, bg=COLORS['bg_tertiary'])
        period_frame.pack(side=tk.LEFT, padx=10)
        
        periods = [('Week', 'week'), ('Month', 'month'), ('Year', 'year')]
        self.period_buttons = {}
        
        for text, value in periods:
            btn = tk.Button(
                period_frame,
                text=text,
                font=FONTS['body'],
                bg=COLORS['primary'] if value == self.current_period else COLORS['bg_tertiary'],
                fg=COLORS['text_white'] if value == self.current_period else COLORS['text_secondary'],
                relief=tk.FLAT,
                padx=15,
                pady=8,
                cursor='hand2',
                command=lambda v=value: self.change_period(v)
            )
            btn.pack(side=tk.LEFT, padx=1)
            self.period_buttons[value] = btn
        
        # Chart type selector
        chart_frame = tk.Frame(right, bg=COLORS['bg_tertiary'])
        chart_frame.pack(side=tk.LEFT, padx=10)
        
        charts = [('🥧', 'pie'), ('📊', 'bar'), ('📈', 'line')]
        self.chart_buttons = {}
        
        for icon, value in charts:
            btn = tk.Button(
                chart_frame,
                text=icon,
                font=FONTS['body_large'],
                bg=COLORS['primary'] if value == self.current_chart else COLORS['bg_tertiary'],
                fg=COLORS['text_white'] if value == self.current_chart else COLORS['text_secondary'],
                relief=tk.FLAT,
                padx=12,
                pady=5,
                cursor='hand2',
                command=lambda v=value: self.change_chart(v)
            )
            btn.pack(side=tk.LEFT, padx=1)
            self.chart_buttons[value] = btn
        
        # Export button
        export_btn = tk.Button(
            right,
            text="📥 Export",
            font=FONTS['button'],
            bg=COLORS['secondary'],
            fg=COLORS['text_white'],
            relief=tk.FLAT,
            padx=20,
            pady=10,
            cursor='hand2',
            command=self.export_report
        )
        export_btn.pack(side=tk.LEFT, padx=(10, 0))
        export_btn.bind('<Enter>', lambda e: export_btn.config(bg=COLORS['secondary_dark']))
        export_btn.bind('<Leave>', lambda e: export_btn.config(bg=COLORS['secondary']))
    
    def change_period(self, period):
        """Change time period"""
        self.current_period = period
        
        # Update button styles
        for p, btn in self.period_buttons.items():
            if p == period:
                btn.config(bg=COLORS['primary'], fg=COLORS['text_white'])
            else:
                btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'])
        
        self.load_data()
    
    def change_chart(self, chart_type):
        """Change chart type"""
        self.current_chart = chart_type
        
        # Update button styles
        for c, btn in self.chart_buttons.items():
            if c == chart_type:
                btn.config(bg=COLORS['primary'], fg=COLORS['text_white'])
            else:
                btn.config(bg=COLORS['bg_tertiary'], fg=COLORS['text_secondary'])
        
        self.create_main_chart()
    
    def create_summary_cards(self, stats):
        """Create summary stat cards"""
        # Clear existing
        for widget in self.summary_frame.winfo_children():
            widget.destroy()
        
        # Configure grid
        for i in range(5):
            self.summary_frame.columnconfigure(i, weight=1)
        
        cards = [
            ("Total Spent", format_currency(stats['total']), "💰", COLORS['primary']),
            ("Transactions", str(stats['count']), "🧾", COLORS['secondary']),
            ("Average", format_currency(stats['avg']), "📊", COLORS['accent']),
            ("Highest", format_currency(stats['max']), "⬆️", COLORS['error']),
            ("Lowest", format_currency(stats['min']), "⬇️", COLORS['success'])
        ]
        
        for i, (title, value, icon, color) in enumerate(cards):
            card = tk.Frame(self.summary_frame, bg=COLORS['card_bg'])
            card.grid(row=0, column=i, sticky='nsew', padx=6, pady=5)
            
            inner = tk.Frame(card, bg=COLORS['card_bg'])
            inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
            
            top = tk.Frame(inner, bg=COLORS['card_bg'])
            top.pack(fill=tk.X)
            
            tk.Label(
                top,
                text=icon,
                font=('Segoe UI', 18),
                bg=COLORS['card_bg']
            ).pack(side=tk.LEFT)
            
            tk.Label(
                top,
                text=title,
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_secondary']
            ).pack(side=tk.RIGHT)
            
            tk.Label(
                inner,
                text=value,
                font=FONTS['heading_small'],
                bg=COLORS['card_bg'],
                fg=color
            ).pack(anchor='w', pady=(10, 0))
    
    def create_main_chart(self):
        """Create main chart based on selected type"""
        # Clear existing
        for widget in self.main_chart_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = tk.Frame(self.main_chart_frame, bg=COLORS['card_bg'])
        header.pack(fill=tk.X, padx=25, pady=(20, 10))
        
        chart_titles = {
            'pie': '🥧 Spending Distribution',
            'bar': '📊 Category Comparison',
            'line': '📈 Spending Trend'
        }
        
        tk.Label(
            header,
            text=chart_titles[self.current_chart],
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(side=tk.LEFT)
        
        if not MATPLOTLIB_AVAILABLE:
            self.show_no_matplotlib()
            return
        
        if not hasattr(self, 'category_data') or not self.category_data:
            self.show_no_data_chart("No data available")
            return
        
        if self.current_chart == 'pie':
            self.create_pie_chart()
        elif self.current_chart == 'bar':
            self.create_bar_chart()
        else:
            self.create_line_chart()
    
    def create_pie_chart(self):
        """Create detailed pie chart"""
        data = [(c['category_name'].split(' ')[-1][:12], float(c['total']), c.get('color', CHART_COLORS[0])) 
                for c in self.category_data if float(c['total']) > 0][:10]
        
        if not data:
            self.show_no_data_chart("No expenses in this period")
            return
        
        fig = Figure(figsize=(6, 4.5), dpi=100, facecolor=COLORS['card_bg'])
        ax = fig.add_subplot(111)
        
        labels = [d[0] for d in data]
        sizes = [d[1] for d in data]
        colors = [d[2] for d in data]
        
        explode = [0.03] * len(data)
        
        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            colors=colors,
            autopct='%1.1f%%',
            startangle=90,
            explode=explode,
            textprops={'fontsize': 9},
            pctdistance=0.75,
            shadow=True
        )
        
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontsize(8)
            autotext.set_fontweight('bold')
        
        ax.axis('equal')
        
        # Add legend
        ax.legend(wedges, labels, title="Categories", loc="center left", 
                  bbox_to_anchor=(1, 0, 0.5, 1), fontsize=8)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.main_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=20, pady=15, fill=tk.BOTH, expand=True)
    
    def create_bar_chart(self):
        """Create horizontal bar chart"""
        data = [(c['category_name'].split(' ')[-1][:12], float(c['total']), c.get('color', CHART_COLORS[0])) 
                for c in self.category_data if float(c['total']) > 0][:8]
        
        if not data:
            self.show_no_data_chart("No expenses in this period")
            return
        
        fig = Figure(figsize=(6, 4.5), dpi=100, facecolor=COLORS['card_bg'])
        ax = fig.add_subplot(111)
        
        labels = [d[0] for d in data]
        values = [d[1] for d in data]
        colors = [d[2] for d in data]
        
        y_pos = range(len(labels))
        bars = ax.barh(y_pos, values, color=colors, height=0.6, edgecolor='white')
        
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=9)
        ax.invert_yaxis()
        
        ax.set_facecolor(COLORS['card_bg'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.tick_params(colors=COLORS['text_secondary'], left=False)
        
        # Add value labels
        for bar, value in zip(bars, values):
            width = bar.get_width()
            ax.text(width + max(values) * 0.02, bar.get_y() + bar.get_height()/2,
                   f'₹{value:,.0f}', ha='left', va='center', fontsize=8,
                   color=COLORS['text_secondary'])
        
        ax.set_xlim(0, max(values) * 1.3)
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.main_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=20, pady=15, fill=tk.BOTH, expand=True)
    
    def create_line_chart(self):
        """Create line chart for trends"""
        if not hasattr(self, 'daily_data') or not self.daily_data:
            self.show_no_data_chart("No daily data available")
            return
        
        fig = Figure(figsize=(6, 4.5), dpi=100, facecolor=COLORS['card_bg'])
        ax = fig.add_subplot(111)
        
        # Handle both 'date' and 'day' keys from different data sources
        dates = [d.get('date') or d.get('day') or '' for d in self.daily_data]
        values = [float(d['total']) for d in self.daily_data]
        
        # Fill area under line
        ax.fill_between(range(len(dates)), values, alpha=0.3, color=COLORS['primary'])
        ax.plot(range(len(dates)), values, color=COLORS['primary'], linewidth=2.5, 
                marker='o', markersize=4, markerfacecolor=COLORS['text_white'],
                markeredgecolor=COLORS['primary'])
        
        ax.set_facecolor(COLORS['card_bg'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.tick_params(colors=COLORS['text_secondary'], labelsize=8)
        
        # Format x-axis
        step = max(1, len(dates) // 7)
        ax.set_xticks(range(0, len(dates), step))
        ax.set_xticklabels([dates[i][-5:] for i in range(0, len(dates), step)], rotation=45)
        
        ax.yaxis.set_major_formatter(plt.FuncFormatter(
            lambda x, p: f'₹{x/1000:.1f}k' if x >= 1000 else f'₹{x:.0f}'))
        
        ax.grid(True, axis='y', linestyle='--', alpha=0.3)
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.main_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=20, pady=15, fill=tk.BOTH, expand=True)
    
    def create_category_breakdown(self):
        """Create category breakdown panel"""
        # Clear existing
        for widget in self.breakdown_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = tk.Frame(self.breakdown_frame, bg=COLORS['card_bg'])
        header.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(
            header,
            text="📋 Category Breakdown",
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(anchor='w')
        
        if not hasattr(self, 'category_data') or not self.category_data:
            tk.Label(
                self.breakdown_frame,
                text="No data available",
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_secondary']
            ).pack(pady=20)
            return
        
        total = sum(float(c['total']) for c in self.category_data)
        
        for cat in self.category_data[:8]:
            if float(cat['total']) > 0:
                self.create_breakdown_row(cat, total)
    
    def create_breakdown_row(self, category, total):
        """Create category breakdown row"""
        row = tk.Frame(self.breakdown_frame, bg=COLORS['card_bg'])
        row.pack(fill=tk.X, padx=20, pady=8)
        
        percentage = (float(category['total']) / total * 100) if total > 0 else 0
        
        # Color indicator
        color_bar = tk.Frame(row, bg=category.get('color', COLORS['primary']), width=4)
        color_bar.pack(side=tk.LEFT, fill=tk.Y, pady=2)
        
        # Info
        info = tk.Frame(row, bg=COLORS['card_bg'])
        info.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        
        name_row = tk.Frame(info, bg=COLORS['card_bg'])
        name_row.pack(fill=tk.X)
        
        icon = category.get('icon', '📦')
        name = category['category_name'].split(' ', 1)[-1][:15] if ' ' in category['category_name'] else category['category_name'][:15]
        
        tk.Label(
            name_row,
            text=f"{icon} {name}",
            font=FONTS['body_medium'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(side=tk.LEFT)
        
        tk.Label(
            name_row,
            text=f"{percentage:.1f}%",
            font=FONTS['body'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_secondary']
        ).pack(side=tk.RIGHT)
        
        # Progress bar
        progress_frame = tk.Frame(info, bg=COLORS['bg_tertiary'], height=6)
        progress_frame.pack(fill=tk.X, pady=(5, 0))
        
        progress = tk.Frame(progress_frame, bg=category.get('color', COLORS['primary']), height=6)
        progress.place(relwidth=percentage/100, relheight=1)
        
        # Amount
        tk.Label(
            info,
            text=format_currency(category['total']),
            font=FONTS['body'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_secondary']
        ).pack(anchor='w', pady=(5, 0))
    
    def create_daily_trend_chart(self):
        """Create daily trend bar chart"""
        # Clear existing
        for widget in self.daily_chart_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = tk.Frame(self.daily_chart_frame, bg=COLORS['card_bg'])
        header.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        tk.Label(
            header,
            text="📅 Daily Spending",
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(side=tk.LEFT)
        
        if not MATPLOTLIB_AVAILABLE or not hasattr(self, 'daily_data') or not self.daily_data:
            self.show_simple_message(self.daily_chart_frame, "No daily data available")
            return
        
        fig = Figure(figsize=(5, 3), dpi=100, facecolor=COLORS['card_bg'])
        ax = fig.add_subplot(111)
        
        dates = [str(d.get('day', d.get('date', '')))[-2:] for d in self.daily_data[-14:]]  # Last 14 days
        values = [float(d['total']) for d in self.daily_data[-14:]]
        
        colors = [COLORS['primary'] if v < sum(values)/len(values) else COLORS['warning'] for v in values]
        
        bars = ax.bar(range(len(dates)), values, color=colors, width=0.7, edgecolor='white')
        
        ax.set_xticks(range(len(dates)))
        ax.set_xticklabels(dates, fontsize=7)
        ax.set_facecolor(COLORS['card_bg'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.tick_params(colors=COLORS['text_secondary'], labelsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(
            lambda x, p: f'₹{x/1000:.0f}k' if x >= 1000 else f'₹{x:.0f}'))
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.daily_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=10, pady=(0, 15))
    
    def create_monthly_comparison_chart(self):
        """Create monthly comparison chart"""
        # Clear existing
        for widget in self.monthly_chart_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = tk.Frame(self.monthly_chart_frame, bg=COLORS['card_bg'])
        header.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        tk.Label(
            header,
            text="📈 Monthly Comparison",
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(side=tk.LEFT)
        
        if not MATPLOTLIB_AVAILABLE or not hasattr(self, 'monthly_data') or not self.monthly_data:
            self.show_simple_message(self.monthly_chart_frame, "No monthly data available")
            return
        
        fig = Figure(figsize=(5, 3), dpi=100, facecolor=COLORS['card_bg'])
        ax = fig.add_subplot(111)
        
        months = [get_month_short_name(d['month']) for d in self.monthly_data]
        values = [float(d['total']) for d in self.monthly_data]
        
        # Gradient colors
        colors = [COLORS['primary'] if i == len(values)-1 else COLORS['primary_light'] 
                  for i in range(len(values))]
        
        bars = ax.bar(range(len(months)), values, color=colors, width=0.6, edgecolor='white')
        
        ax.set_xticks(range(len(months)))
        ax.set_xticklabels(months, fontsize=8)
        ax.set_facecolor(COLORS['card_bg'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.tick_params(colors=COLORS['text_secondary'], labelsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(
            lambda x, p: f'₹{x/1000:.0f}k' if x >= 1000 else f'₹{x:.0f}'))
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.monthly_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=10, pady=(0, 15))
    
    def create_expense_table(self):
        """Create expense data table"""
        # Clear existing
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = tk.Frame(self.table_frame, bg=COLORS['card_bg'])
        header.pack(fill=tk.X, padx=25, pady=(20, 15))
        
        tk.Label(
            header,
            text="📝 Recent Transactions",
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(side=tk.LEFT)
        
        if not hasattr(self, 'expenses') or not self.expenses:
            tk.Label(
                self.table_frame,
                text="No transactions in this period",
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_secondary']
            ).pack(pady=20)
            return
        
        # Table headers
        columns = tk.Frame(self.table_frame, bg=COLORS['bg_tertiary'])
        columns.pack(fill=tk.X, padx=25)
        
        headers = [("Date", 100), ("Category", 150), ("Description", 250), ("Amount", 120)]
        for text, width in headers:
            tk.Label(
                columns,
                text=text,
                font=FONTS['body_medium'],
                bg=COLORS['bg_tertiary'],
                fg=COLORS['text_secondary'],
                width=width//10,
                anchor='w'
            ).pack(side=tk.LEFT, padx=10, pady=8)
        
        # Table rows
        for expense in self.expenses[:15]:
            row = tk.Frame(self.table_frame, bg=COLORS['card_bg'])
            row.pack(fill=tk.X, padx=25)
            
            # Date
            tk.Label(
                row,
                text=expense.expense_date or "",
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_primary'],
                width=10,
                anchor='w'
            ).pack(side=tk.LEFT, padx=10, pady=8)
            
            # Category
            cat_frame = tk.Frame(row, bg=COLORS['card_bg'])
            cat_frame.pack(side=tk.LEFT, padx=10)
            
            cat_name = expense.category_name.split(' ', 1)[-1][:15] if expense.category_name else "Uncategorized"
            
            tk.Label(
                cat_frame,
                text=f"{expense.category_icon or '📦'} {cat_name}",
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_primary'],
                anchor='w',
                width=15
            ).pack(side=tk.LEFT)
            
            # Description
            tk.Label(
                row,
                text=(expense.description or "")[:30] + ("..." if len(expense.description or "") > 30 else ""),
                font=FONTS['body'],
                bg=COLORS['card_bg'],
                fg=COLORS['text_secondary'],
                width=25,
                anchor='w'
            ).pack(side=tk.LEFT, padx=10)
            
            # Amount
            tk.Label(
                row,
                text=format_currency(expense.amount),
                font=FONTS['body_medium'],
                bg=COLORS['card_bg'],
                fg=COLORS['error'],
                width=12,
                anchor='e'
            ).pack(side=tk.LEFT, padx=10)
            
            # Separator
            tk.Frame(self.table_frame, bg=COLORS['border'], height=1).pack(fill=tk.X, padx=25)
        
        tk.Frame(self.table_frame, bg=COLORS['card_bg'], height=20).pack()
    
    def show_no_matplotlib(self):
        """Show message when matplotlib is not available"""
        container = tk.Frame(self.main_chart_frame, bg=COLORS['card_bg'])
        container.pack(fill=tk.BOTH, expand=True, pady=40)
        
        tk.Label(
            container,
            text="📊",
            font=('Segoe UI', 48),
            bg=COLORS['card_bg']
        ).pack()
        
        tk.Label(
            container,
            text="Charts require matplotlib",
            font=FONTS['subheading'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_primary']
        ).pack(pady=(15, 5))
        
        tk.Label(
            container,
            text="Run: pip install matplotlib",
            font=FONTS['body'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_secondary']
        ).pack()
    
    def show_no_data_chart(self, message):
        """Show no data message in chart area"""
        container = tk.Frame(self.main_chart_frame, bg=COLORS['card_bg'])
        container.pack(fill=tk.BOTH, expand=True, pady=40)
        
        tk.Label(
            container,
            text="📭",
            font=('Segoe UI', 48),
            bg=COLORS['card_bg']
        ).pack()
        
        tk.Label(
            container,
            text=message,
            font=FONTS['body_medium'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_secondary']
        ).pack(pady=(15, 0))
    
    def show_simple_message(self, parent, message):
        """Show simple message"""
        tk.Label(
            parent,
            text=message,
            font=FONTS['body'],
            bg=COLORS['card_bg'],
            fg=COLORS['text_secondary']
        ).pack(pady=20)
    
    def export_report(self):
        """Export report to Excel"""
        if not hasattr(self, 'expenses') or not self.expenses:
            messagebox.showinfo("Export", "No data to export")
            return
        
        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "Excel export requires openpyxl library.\nInstall it with: pip install openpyxl")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"expense_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if filepath:
            try:
                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "Expense Report"
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                currency_font = Font(bold=True, color="10B981")
                
                # Add title
                ws.merge_cells('A1:E1')
                ws['A1'] = f"📊 Expense Report - {datetime.now().strftime('%B %Y')}"
                ws['A1'].font = Font(bold=True, size=16, color="667EEA")
                ws['A1'].alignment = Alignment(horizontal="center")
                
                # Add summary section
                ws['A3'] = "Report Summary"
                ws['A3'].font = Font(bold=True, size=12)
                
                total_amount = sum(float(e.amount) for e in self.expenses)
                ws['A4'] = f"Total Expenses: ₹{total_amount:,.2f}"
                ws['A5'] = f"Number of Transactions: {len(self.expenses)}"
                ws['A6'] = f"Average Expense: ₹{(total_amount / len(self.expenses)):,.2f}" if self.expenses else "₹0.00"
                ws['A7'] = f"Report Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
                
                # Headers for expense list (starting from row 9)
                headers = ["#", "Date", "Category", "Description", "Amount (₹)"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=9, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Add expense data
                for row_idx, expense in enumerate(self.expenses, 10):
                    cat_name = expense.category_name if expense.category_name else "Uncategorized"
                    # Remove emoji from category for cleaner Excel display
                    if cat_name and len(cat_name) > 2 and ord(cat_name[0]) > 127:
                        cat_name = cat_name.split(' ', 1)[-1] if ' ' in cat_name else cat_name
                    
                    ws.cell(row=row_idx, column=1, value=row_idx - 9).border = thin_border
                    ws.cell(row=row_idx, column=2, value=str(expense.expense_date)).border = thin_border
                    ws.cell(row=row_idx, column=3, value=cat_name).border = thin_border
                    ws.cell(row=row_idx, column=4, value=expense.description or "").border = thin_border
                    amount_cell = ws.cell(row=row_idx, column=5, value=float(expense.amount))
                    amount_cell.border = thin_border
                    amount_cell.number_format = '₹#,##0.00'
                
                # Add total row
                total_row = 10 + len(self.expenses)
                ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
                total_cell = ws.cell(row=total_row, column=5, value=total_amount)
                total_cell.font = Font(bold=True, color="667EEA")
                total_cell.number_format = '₹#,##0.00'
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 35
                ws.column_dimensions['E'].width = 15
                
                # Add category summary sheet
                ws2 = wb.create_sheet(title="Category Summary")
                ws2['A1'] = "Category-wise Summary"
                ws2['A1'].font = Font(bold=True, size=14, color="667EEA")
                
                # Category headers
                cat_headers = ["Category", "Total Amount (₹)", "Transactions", "Percentage"]
                for col, header in enumerate(cat_headers, 1):
                    cell = ws2.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Calculate category totals
                category_totals = {}
                for expense in self.expenses:
                    cat = expense.category_name or "Uncategorized"
                    if cat and len(cat) > 2 and ord(cat[0]) > 127:
                        cat = cat.split(' ', 1)[-1] if ' ' in cat else cat
                    if cat not in category_totals:
                        category_totals[cat] = {'total': 0, 'count': 0}
                    category_totals[cat]['total'] += float(expense.amount)
                    category_totals[cat]['count'] += 1
                
                row = 4
                for cat_name, data in sorted(category_totals.items(), key=lambda x: x[1]['total'], reverse=True):
                    pct = (data['total'] / total_amount * 100) if total_amount > 0 else 0
                    ws2.cell(row=row, column=1, value=cat_name).border = thin_border
                    ws2.cell(row=row, column=2, value=data['total']).border = thin_border
                    ws2.cell(row=row, column=2).number_format = '₹#,##0.00'
                    ws2.cell(row=row, column=3, value=data['count']).border = thin_border
                    ws2.cell(row=row, column=4, value=f"{pct:.1f}%").border = thin_border
                    row += 1
                
                # Adjust category sheet column widths
                ws2.column_dimensions['A'].width = 25
                ws2.column_dimensions['B'].width = 18
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 12
                
                # Save the workbook
                wb.save(filepath)
                
                messagebox.showinfo("✅ Success", f"Report exported successfully!\n\nFile: {filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def load_data(self):
        """Load report data"""
        # Get date range based on period
        today = datetime.now()
        
        if self.current_period == 'week':
            start_date = (today - timedelta(days=7)).strftime('%Y-%m-%d')
        elif self.current_period == 'month':
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
        else:  # year
            start_date = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        
        end_date = today.strftime('%Y-%m-%d')
        
        # Load data
        data = ExpenseController.get_report_data(self.user.user_id, start_date, end_date)
        
        self.category_data = data.get('category_totals', [])
        self.daily_data = data.get('daily_trend', [])
        self.monthly_data = data.get('monthly_trend', [])
        self.expenses = data.get('expenses', [])
        
        # Calculate stats
        total = sum(float(c['total']) for c in self.category_data) if self.category_data else 0
        count = len(self.expenses)
        amounts = [float(e.amount) for e in self.expenses] if self.expenses else [0]
        
        stats = {
            'total': total,
            'count': count,
            'avg': total / count if count > 0 else 0,
            'max': max(amounts) if amounts else 0,
            'min': min(amounts) if amounts else 0
        }
        
        # Create UI elements
        self.create_summary_cards(stats)
        self.create_main_chart()
        self.create_category_breakdown()
        self.create_daily_trend_chart()
        self.create_monthly_comparison_chart()
        self.create_expense_table()
    
    def refresh(self):
        """Refresh report data"""
        self.load_data()
